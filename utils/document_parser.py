"""
Document parsing utilities for handling various file formats
"""
import os
import PyPDF2
import pandas as pd
from docx import Document
import openpyxl
import traceback
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def parse_document(file_path):
    """
    Parse a document based on its file extension

    Args:
        file_path (str): Path to the document

    Returns:
        str: Extracted text content from the document
    """
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    try:
        if ext == '.pdf':
            return parse_pdf(file_path)
        elif ext == '.docx':
            return parse_docx(file_path)
        elif ext == '.xlsx' or ext == '.xls':
            return parse_excel(file_path)
        elif ext == '.csv':
            return parse_csv(file_path)
        elif ext in ['.txt', '.md', '.html', '.htm']:
            return parse_text(file_path)
        else:
            # Default to treating as text
            try:
                return parse_text(file_path)
            except:
                return f"Could not parse file with extension {ext}. Supported formats are PDF, DOCX, XLSX, XLS, CSV, and text files."
    except Exception as e:
        logger.error(f"Error parsing document {file_path}: {str(e)}")
        logger.error(traceback.format_exc())
        return f"Error parsing document: {str(e)}"


def parse_pdf(file_path):
    """
    Extract text from a PDF file

    Args:
        file_path (str): Path to the PDF file

    Returns:
        str: Extracted text content
    """
    content = ""
    with open(file_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        num_pages = len(reader.pages)

        for page_num in range(num_pages):
            page = reader.pages[page_num]
            content += page.extract_text() + "\n\n"

    return content


def parse_docx(file_path):
    """
    Extract text from a Word document

    Args:
        file_path (str): Path to the DOCX file

    Returns:
        str: Extracted text content
    """
    document = Document(file_path)
    content = []

    for para in document.paragraphs:
        content.append(para.text)

    return '\n'.join(content)


def parse_excel(file_path):
    """
    Extract data from an Excel document

    Args:
        file_path (str): Path to the Excel file

    Returns:
        str: Extracted data as formatted text
    """
    try:
        # Try using pandas first
        excel_data = pd.read_excel(file_path, sheet_name=None)
        content = []

        for sheet_name, df in excel_data.items():
            content.append(f"Sheet: {sheet_name}")
            content.append(df.to_string(index=False))
            content.append("\n")

        return '\n'.join(content)
    except Exception as e:
        logger.warning(f"Could not parse Excel with pandas: {str(e)}. Trying with openpyxl...")

        # Fallback to openpyxl
        content = []
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content.append(f"Sheet: {sheet_name}")

            for row in sheet.iter_rows(values_only=True):
                row_values = [str(val) if val is not None else "" for val in row]
                content.append('\t'.join(row_values))

            content.append("\n")

        return '\n'.join(content)


def parse_csv(file_path):
    """
    Extract data from a CSV file

    Args:
        file_path (str): Path to the CSV file

    Returns:
        str: Extracted data as formatted text
    """
    try:
        df = pd.read_csv(file_path)
        return df.to_string(index=False)
    except Exception as e:
        logger.error(f"Error parsing CSV: {str(e)}")

        # Fallback to basic text parsing
        return parse_text(file_path)


def parse_text(file_path):
    """
    Extract content from a text file

    Args:
        file_path (str): Path to the text file

    Returns:
        str: Extracted text content
    """
    with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
        return file.read()